#!/usr/bin/env python3
"""
Batch analysis for stripping-cell Excel files.

The script expects a root directory with one first-level folder per sample, for example:

root_dir/
    Sample_A/
        repeat_1/*.xlsx
        repeat_2/*.xlsx
    Sample_B/
        repeat_1/*.xlsx

Main outputs are written to:

root_dir/stripping_outputs/
    figures/                 one Voltage-Capacity plot per sample
    plot_data/               one CSV per sample/repeat/file used for plotting
    summary/                 summary tables and color map

Summary metric definitions
--------------------------
Nucleation (ohm): magnitude of the first negative local voltage valley. The column
    name follows the user's requested output name; the numerical calculation is in mV.
Plateau (mV): magnitude of voltage near half of the step-sheet capacity, matching the
    logic in the original summary script.
Overp. (mV): Nucleation - Plateau, when both values are numeric. This represents the
    extra initial nucleation barrier relative to the plateau.
Cap. (mAh/cm2): step-sheet H3 capacity divided by electrode area.

Notes
-----
- By default, plotting capacity is normalized by electrode area, so the x-axis is
  Capacity (mAh/cm^2). To reproduce the old plotting script's test-sheet divisor,
  use: --normalization test-dchg
- Each sample is assigned one color, and all repeats/files within that sample are
  plotted with the same solid line style.
- If --xlim is omitted, each sample plot automatically uses an X-axis maximum of
  2.0 or 7.5 based on the valid maximum plotting capacity. Files above the short
  threshold are marked as short in the summary and do not stretch the axis.
- Documentation and comments are intentionally written in English.
"""

from __future__ import annotations

import argparse
import math
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.colors import to_hex
from matplotlib.ticker import AutoMinorLocator
from openpyxl import load_workbook


TABLEAU_10 = [
    "#4E79A7",  # blue
    "#F28E2B",  # orange
    "#59A14F",  # green
    "#E15759",  # red
    "#76B7B2",  # teal
    "#EDC948",  # yellow
    "#B07AA1",  # purple
    "#FF9DA7",  # pink
    "#9C755F",  # brown
    "#BAB0AC",  # gray
]

@dataclass
class FileInfo:
    path: Path
    sample: str
    repeat: str


@dataclass
class MetaData:
    displayed_file_name: str
    time: str
    operator: str
    step_capacity_mAh: Optional[float]
    areal_capacity_mAh_cm2: object
    status: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Batch process stripping-cell Excel files by sample folder."
    )
    parser.add_argument(
        "root_dir",
        type=Path,
        help="Root directory containing one first-level folder per sample.",
    )
    parser.add_argument(
        "--output-dir-name",
        default="stripping_outputs",
        help="Name of the output folder created inside root_dir.",
    )
    parser.add_argument(
        "--area",
        type=float,
        default=1.27,
        help="Electrode area in cm^2 used for capacity normalization. Default: 1.27.",
    )
    parser.add_argument(
        "--operator",
        default="Sravani",
        help="Operator name written to the summary when it is not available in the file.",
    )
    parser.add_argument(
        "--normalization",
        choices=["area", "test-dchg", "none"],
        default="area",
        help=(
            "How to normalize plotting capacity: 'area' gives mAh/cm^2; "
            "'test-dchg' reproduces the old test-sheet divisor; 'none' keeps raw mAh."
        ),
    )
    parser.add_argument(
        "--step-type",
        default="CC DChg",
        help="Step Type selected from the record sheet for plotting. Default: CC DChg.",
    )
    parser.add_argument(
        "--valley-window",
        type=int,
        default=3,
        help="Neighbor window used to detect the first voltage valley. Default: 3.",
    )
    parser.add_argument(
        "--xlim",
        type=float,
        nargs=2,
        default=None,
        metavar=("XMIN", "XMAX"),
        help=(
            "Manual X-axis limits for all sample plots. If omitted, the script "
            "automatically uses either --small-capacity-limit or --large-capacity-limit."
        ),
    )
    parser.add_argument(
        "--x-min",
        type=float,
        default=-0.5,
        help="Lower X-axis limit used when --xlim is omitted. Default: -0.5.",
    )
    parser.add_argument(
        "--small-capacity-limit",
        type=float,
        default=2.0,
        help=(
            "Automatic X-axis upper limit for samples whose valid maximum plotting "
            "capacity is at or below this value. Default: 2.0."
        ),
    )
    parser.add_argument(
        "--large-capacity-limit",
        type=float,
        default=7.5,
        help=(
            "Automatic X-axis upper limit for samples whose valid maximum plotting "
            "capacity is above --small-capacity-limit. Default: 7.5."
        ),
    )
    parser.add_argument(
        "--short-capacity-threshold",
        type=float,
        default=None,
        help=(
            "Files with maximum plotting capacity above this threshold are marked "
            "as short in the summary. Default: same as --large-capacity-limit."
        ),
    )
    parser.add_argument(
        "--ylim",
        type=float,
        nargs=2,
        default=(-1.0, 0.2),
        metavar=("YMIN", "YMAX"),
        help="Y-axis limits for sample plots. Default: -1.0 0.2.",
    )
    parser.add_argument(
        "--dpi",
        type=int,
        default=300,
        help="DPI for saved figures. Default: 300.",
    )
    parser.add_argument(
        "--show-legend",
        action="store_true",
        help="Show a repeat/file legend on each sample plot.",
    )
    return parser.parse_args()


def set_plot_style() -> None:
    """Apply consistent scientific plotting style."""
    plt.rcParams["font.family"] = "sans-serif"
    plt.rcParams["font.sans-serif"] = ["Arial", "Helvetica", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["pdf.fonttype"] = 42
    plt.rcParams["ps.fonttype"] = 42


def safe_name(text: object) -> str:
    """Return a filesystem-safe version of a sample/repeat/file name."""
    text = str(text).strip()
    text = re.sub(r"[\\/:*?\"<>|]+", "_", text)
    text = re.sub(r"\s+", "_", text)
    return text or "unnamed"


def normalized_col_name(name: object) -> str:
    """Normalize Excel column names for tolerant matching."""
    return re.sub(r"[^a-z0-9]+", "", str(name).lower())


def find_column(
    df: pd.DataFrame,
    candidates: Iterable[str],
    fallback_index: Optional[int] = None,
) -> Optional[str]:
    """Find a DataFrame column by tolerant name matching, with optional index fallback."""
    normalized = {normalized_col_name(col): col for col in df.columns}
    for candidate in candidates:
        key = normalized_col_name(candidate)
        if key in normalized:
            return normalized[key]
    if fallback_index is not None and 0 <= fallback_index < len(df.columns):
        return str(df.columns[fallback_index])
    return None


def collect_excel_files(root_dir: Path, output_dir_name: str) -> list[FileInfo]:
    """Recursively collect Excel files and infer sample/repeat from folder structure."""
    files: list[FileInfo] = []
    output_dir = (root_dir / output_dir_name).resolve()

    for file_path in sorted(root_dir.rglob("*.xlsx")):
        if file_path.name.startswith("~$"):
            continue
        if "summary" in file_path.name.lower():
            continue
        if output_dir in file_path.resolve().parents:
            continue

        rel_parent = file_path.parent.relative_to(root_dir)
        parts = rel_parent.parts

        if len(parts) == 0:
            sample = root_dir.name
            repeat = file_path.stem
        else:
            sample = parts[0]
            repeat = " / ".join(parts[1:]) if len(parts) > 1 else file_path.stem

        files.append(FileInfo(path=file_path, sample=sample, repeat=repeat))

    return files


def build_sample_color_map(samples: list[str]) -> dict[str, str]:
    """Assign stable, publication-friendly colors to sorted sample names."""
    samples = sorted(samples)
    n = len(samples)

    colors: list[str] = []
    if n <= len(TABLEAU_10):
        colors = TABLEAU_10[:n]
    else:
        colors.extend(TABLEAU_10)
        tab20 = [to_hex(c) for c in plt.get_cmap("tab20").colors]
        for color in tab20:
            if len(colors) >= n:
                break
            if color.upper() not in {c.upper() for c in colors}:
                colors.append(color)

        if len(colors) < n:
            extra = n - len(colors)
            # Evenly spaced hue colors are used only when there are more samples than
            # the fixed categorical palettes can cover.
            colors.extend(to_hex(c) for c in plt.get_cmap("hsv")(np.linspace(0, 0.85, extra)))

    return dict(zip(samples, colors[:n]))


def read_metadata(file_path: Path, area: float, default_operator: str) -> MetaData:
    """Read file-level metadata from unit and step sheets."""
    displayed_file_name = file_path.name
    time_value: object = "N/A"
    operator = default_operator
    step_capacity_mAh: Optional[float] = None
    areal_capacity: object = "N/A"
    status = "ok"

    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)

        if "unit" in wb.sheetnames:
            unit_ws = wb["unit"]
            if unit_ws["A1"].value:
                displayed_file_name = str(unit_ws["A1"].value)

            raw_time = unit_ws["C3"].value
            if isinstance(raw_time, datetime):
                time_value = raw_time.date().isoformat()
            elif raw_time is not None:
                time_value = str(raw_time).split()[0]

        if "step" in wb.sheetnames:
            step_ws = wb["step"]
            h3 = step_ws["H3"].value
            if isinstance(h3, (int, float)) and not pd.isna(h3) and h3 != 0:
                step_capacity_mAh = float(h3)
                areal_capacity = step_capacity_mAh / area
            else:
                status = "cell shorted"
                areal_capacity = "cell shorted"
        else:
            status = "missing step sheet"

        wb.close()

    except Exception as exc:
        status = f"metadata error: {exc}"
        areal_capacity = "file broken"

    return MetaData(
        displayed_file_name=displayed_file_name,
        time=str(time_value),
        operator=operator,
        step_capacity_mAh=step_capacity_mAh,
        areal_capacity_mAh_cm2=areal_capacity,
        status=status,
    )


def get_test_dchg_divisor(file_path: Path) -> float:
    """Reproduce the old plotting script's test-sheet DChg capacity divisor."""
    try:
        test_df = pd.read_excel(file_path, sheet_name="test")
        col = find_column(test_df, ["DChg. Cap.(mAh)", "DChg Cap (mAh)"])
        if col is None:
            return 1.0
        values = pd.to_numeric(test_df[col].iloc[2:], errors="coerce").dropna().reset_index(drop=True)
        if len(values) == 0 or values.iloc[0] == 0:
            return 1.0
        return float(values.iloc[0])
    except Exception:
        return 1.0


def read_record_data(
    file_info: FileInfo,
    metadata: MetaData,
    area: float,
    normalization: str,
    step_type: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Read the record sheet and return two tables:
    1. record_numeric: all rows with numeric voltage and capacity.
    2. plot_data: filtered rows used for plotting.
    """
    record_df = pd.read_excel(file_info.path, sheet_name="record")

    step_col = find_column(record_df, ["Step Type", "StepType"])
    voltage_col = find_column(record_df, ["Voltage(V)", "Voltage (V)", "Voltage"], fallback_index=5)
    capacity_col = find_column(record_df, ["Capacity(mAh)", "Capacity (mAh)", "Capacity"], fallback_index=6)

    if voltage_col is None or capacity_col is None:
        raise ValueError("Could not find voltage/capacity columns in the record sheet.")

    work = record_df.copy()
    work["__Voltage_V"] = pd.to_numeric(work[voltage_col], errors="coerce")
    work["__Capacity_mAh"] = pd.to_numeric(work[capacity_col], errors="coerce")
    work = work.dropna(subset=["__Voltage_V", "__Capacity_mAh"]).copy()

    if step_col is not None:
        work["__Step_Type"] = work[step_col].astype(str).str.strip()
        plot_rows = work[work["__Step_Type"] == step_type].copy()
    else:
        work["__Step_Type"] = ""
        plot_rows = work.copy()

    if plot_rows.empty:
        # Keep the file processable even when the expected step type does not exist.
        plot_rows = work.copy()

    if normalization == "area":
        divisor = area
        x_label = "Capacity (mAh/cm2)"
    elif normalization == "test-dchg":
        divisor = get_test_dchg_divisor(file_info.path)
        x_label = "Normalized capacity by test DChg divisor"
    else:
        divisor = 1.0
        x_label = "Capacity (mAh)"

    if divisor == 0 or pd.isna(divisor):
        divisor = 1.0

    plot_rows["Sample"] = file_info.sample
    plot_rows["Repeat"] = file_info.repeat
    plot_rows["File name"] = metadata.displayed_file_name
    plot_rows["Source path"] = str(file_info.path)
    plot_rows["Voltage (V)"] = plot_rows["__Voltage_V"]
    plot_rows["Capacity (mAh)"] = plot_rows["__Capacity_mAh"]
    plot_rows[x_label] = plot_rows["__Capacity_mAh"] / divisor
    plot_rows["Plot x"] = plot_rows[x_label]
    plot_rows["Plot y"] = plot_rows["__Voltage_V"]

    # Put the most useful columns first, then keep original record columns after them.
    front_cols = [
        "Sample",
        "Repeat",
        "File name",
        "Source path",
        "__Step_Type",
        "Capacity (mAh)",
        x_label,
        "Voltage (V)",
        "Plot x",
        "Plot y",
    ]
    remaining_cols = [c for c in plot_rows.columns if c not in front_cols]
    plot_rows = plot_rows[front_cols + remaining_cols]

    return work, plot_rows


def find_first_valley(values: Iterable[float], window: int = 3) -> Optional[float]:
    """Find the first local negative voltage minimum using a simple neighbor window."""
    clean_values = [float(v) for v in values if isinstance(v, (int, float, np.number)) and not pd.isna(v)]
    if len(clean_values) < 2 * window + 1:
        return None

    for i in range(window, len(clean_values) - window):
        center = clean_values[i]
        left = clean_values[i - window : i]
        right = clean_values[i + 1 : i + 1 + window]
        if center < 0 and all(center < value for value in left + right):
            return center
    return None


def as_float_or_none(value: object) -> Optional[float]:
    """Convert numeric-like values to float, otherwise return None."""
    try:
        if value is None or pd.isna(value):
            return None
        return float(value)
    except Exception:
        return None


def compute_summary_metrics(
    record_numeric: pd.DataFrame,
    metadata: MetaData,
    valley_window: int,
) -> tuple[object, object, object]:
    """Compute nucleation, plateau, and overpotential metrics from voltage/capacity data."""
    if record_numeric.empty:
        return "N/A", "N/A", "N/A"

    voltage = record_numeric["__Voltage_V"].reset_index(drop=True)
    capacity = record_numeric["__Capacity_mAh"].reset_index(drop=True)

    valley = find_first_valley(voltage.tolist(), window=valley_window)
    nucleation_mV: object = -valley * 1000 if valley is not None else "N/A"

    capacity_threshold = metadata.step_capacity_mAh
    if capacity_threshold is None or pd.isna(capacity_threshold):
        max_cap = capacity.max(skipna=True)
        capacity_threshold = float(max_cap) if not pd.isna(max_cap) else None

    plateau_mV: object = "N/A"
    if capacity_threshold is not None and not pd.isna(capacity_threshold):
        half_capacity = 0.5 * float(capacity_threshold)
        below_half = record_numeric[record_numeric["__Capacity_mAh"] < half_capacity]
        if not below_half.empty:
            plateau_voltage = below_half.iloc[-1]["__Voltage_V"]
            plateau_mV = -float(plateau_voltage) * 1000

    nucleation_float = as_float_or_none(nucleation_mV)
    plateau_float = as_float_or_none(plateau_mV)
    if nucleation_float is not None and plateau_float is not None:
        overp_mV: object = nucleation_float - plateau_float
    elif nucleation_float is None and metadata.status == "ok":
        overp_mV = "cell running"
    else:
        overp_mV = "N/A"

    return nucleation_mV, plateau_mV, overp_mV


def append_status(existing_status: str, new_status: str) -> str:
    """Append a status note without duplicating it."""
    existing_status = str(existing_status or "ok")
    if existing_status == "ok":
        return new_status
    status_parts = [part.strip() for part in existing_status.split(";") if part.strip()]
    if new_status not in status_parts:
        status_parts.append(new_status)
    return "; ".join(status_parts)


def max_plot_capacity(plot_df: pd.DataFrame) -> Optional[float]:
    """Return the maximum plotting capacity for one file, if available."""
    if plot_df.empty or "Plot x" not in plot_df.columns:
        return None
    values = pd.to_numeric(plot_df["Plot x"], errors="coerce").dropna()
    if values.empty:
        return None
    return float(values.max())


def determine_sample_xlim(
    sample_rows: list[pd.DataFrame],
    x_min: float,
    small_capacity_limit: float,
    large_capacity_limit: float,
) -> tuple[float, float]:
    """Choose the sample-level X-axis range from valid, non-shorted files."""
    valid_max_values: list[float] = []

    for df in sample_rows:
        if df.empty or df.attrs.get("is_short", False):
            continue
        max_cap = max_plot_capacity(df)
        if max_cap is not None:
            valid_max_values.append(max_cap)

    # If every file in the sample is marked short, use the large limit so the early
    # voltage behavior is still visible while the bad long tail is clipped.
    sample_max = max(valid_max_values) if valid_max_values else large_capacity_limit
    x_max = small_capacity_limit if sample_max <= small_capacity_limit else large_capacity_limit
    return (x_min, x_max)


def save_sample_plot(
    sample: str,
    sample_rows: list[pd.DataFrame],
    sample_color: str,
    figures_dir: Path,
    xlim: tuple[float, float],
    ylim: tuple[float, float],
    dpi: int,
    show_legend: bool,
    x_axis_label: str,
) -> None:
    """Save one plot for one sample, with all repeats/files overlaid."""
    fig, ax = plt.subplots(figsize=(5.2, 4.0))

    for df in sample_rows:
        if df.empty:
            continue
        repeat = str(df["Repeat"].iloc[0])
        file_stem = Path(str(df["Source path"].iloc[0])).stem
        label = f"{repeat} | {file_stem}"
        ax.plot(
            df["Plot x"],
            df["Plot y"],
            color=sample_color,
            linestyle="-",
            linewidth=2.2,
            alpha=0.88,
            label=label,
        )

    ax.set_xlim(*xlim)
    ax.set_ylim(*ylim)
    ax.set_xlabel(x_axis_label, fontsize=16)
    ax.set_ylabel("Voltage (V)", fontsize=16)
    ax.set_title(sample, fontsize=15, pad=10)

    ax.tick_params(axis="both", which="major", direction="in", labelsize=13, length=6, width=1.4)
    ax.tick_params(axis="both", which="minor", direction="in", length=3.5, width=1.0)
    ax.xaxis.set_minor_locator(AutoMinorLocator(2))
    ax.yaxis.set_minor_locator(AutoMinorLocator(2))

    for spine in ax.spines.values():
        spine.set_linewidth(1.4)

    if show_legend:
        ax.legend(fontsize=8.5, frameon=False, loc="best")

    fig.tight_layout()
    png_path = figures_dir / f"{safe_name(sample)}.png"
    pdf_path = figures_dir / f"{safe_name(sample)}.pdf"
    fig.savefig(png_path, dpi=dpi, bbox_inches="tight")
    fig.savefig(pdf_path, bbox_inches="tight")
    plt.close(fig)


def write_outputs(
    output_dir: Path,
    summary_rows: list[dict[str, object]],
    all_plot_rows: list[pd.DataFrame],
    color_map: dict[str, str],
) -> None:
    """Write summary, combined plotting data, and color map files."""
    summary_dir = output_dir / "summary"
    summary_dir.mkdir(parents=True, exist_ok=True)

    summary_columns = [
        "Sample",
        "Repeat",
        "Nucleation (ohm)",
        "Plateau (mV)",
        "Overp. (mV)",
        "Cap. (mAh/cm2)",
        "Time",
        "Operator",
        "File name",
        "Source path",
        "Status",
    ]
    summary_df = pd.DataFrame(summary_rows, columns=summary_columns)
    color_df = pd.DataFrame(
        [{"Sample": sample, "Color": color} for sample, color in sorted(color_map.items())]
    )

    summary_xlsx = summary_dir / "stripping_summary.xlsx"
    with pd.ExcelWriter(summary_xlsx, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="summary")
        color_df.to_excel(writer, index=False, sheet_name="sample_color_map")

    summary_df.to_csv(summary_dir / "stripping_summary.csv", index=False)
    color_df.to_csv(summary_dir / "sample_color_map.csv", index=False)

    if all_plot_rows:
        combined_plot_df = pd.concat(all_plot_rows, ignore_index=True)
        combined_plot_df.to_csv(summary_dir / "all_plot_data.csv", index=False)


def main() -> None:
    args = parse_args()
    root_dir = args.root_dir.expanduser().resolve()
    if not root_dir.exists() or not root_dir.is_dir():
        raise SystemExit(f"Root directory does not exist or is not a directory: {root_dir}")

    set_plot_style()

    output_dir = root_dir / args.output_dir_name
    figures_dir = output_dir / "figures"
    plot_data_dir = output_dir / "plot_data"
    figures_dir.mkdir(parents=True, exist_ok=True)
    plot_data_dir.mkdir(parents=True, exist_ok=True)

    files = collect_excel_files(root_dir, args.output_dir_name)
    if not files:
        raise SystemExit(f"No valid .xlsx files found under: {root_dir}")

    samples = sorted({file_info.sample for file_info in files})
    color_map = build_sample_color_map(samples)

    summary_rows: list[dict[str, object]] = []
    all_plot_rows: list[pd.DataFrame] = []
    sample_to_plot_rows: dict[str, list[pd.DataFrame]] = {sample: [] for sample in samples}
    short_capacity_threshold = (
        args.short_capacity_threshold
        if args.short_capacity_threshold is not None
        else args.large_capacity_limit
    )

    print(f"Found {len(files)} Excel files across {len(samples)} samples.")

    for file_info in files:
        print(f"Processing: sample={file_info.sample} | repeat={file_info.repeat} | file={file_info.path.name}")
        metadata = read_metadata(file_info.path, args.area, args.operator)

        try:
            record_numeric, plot_df = read_record_data(
                file_info=file_info,
                metadata=metadata,
                area=args.area,
                normalization=args.normalization,
                step_type=args.step_type,
            )

            nucleation_mV, plateau_mV, overp_mV = compute_summary_metrics(
                record_numeric=record_numeric,
                metadata=metadata,
                valley_window=args.valley_window,
            )

            file_max_plot_capacity = max_plot_capacity(plot_df)
            is_short = (
                file_max_plot_capacity is not None
                and file_max_plot_capacity > short_capacity_threshold
            )
            plot_df.attrs["is_short"] = is_short
            if is_short:
                metadata.status = append_status(metadata.status, "short")

            sample_plot_dir = plot_data_dir / safe_name(file_info.sample)
            sample_plot_dir.mkdir(parents=True, exist_ok=True)
            csv_name = f"{safe_name(file_info.sample)}__{safe_name(file_info.repeat)}__{safe_name(file_info.path.stem)}__plot_data.csv"
            plot_df.to_csv(sample_plot_dir / csv_name, index=False)

            all_plot_rows.append(plot_df)
            sample_to_plot_rows[file_info.sample].append(plot_df)

        except Exception as exc:
            print(f"  Warning: failed to read record data from {file_info.path.name}: {exc}")
            nucleation_mV = "N/A"
            plateau_mV = "N/A"
            overp_mV = "N/A"
            metadata.status = f"record error: {exc}"

        summary_rows.append(
            {
                "Sample": file_info.sample,
                "Repeat": file_info.repeat,
                "Nucleation (ohm)": nucleation_mV,
                "Plateau (mV)": plateau_mV,
                "Overp. (mV)": overp_mV,
                "Cap. (mAh/cm2)": metadata.areal_capacity_mAh_cm2,
                "Time": metadata.time,
                "Operator": metadata.operator,
                "File name": metadata.displayed_file_name,
                "Source path": str(file_info.path),
                "Status": metadata.status,
            }
        )

    if args.normalization == "area":
        x_axis_label = "Capacity (mAh/cm$^2$)"
    elif args.normalization == "test-dchg":
        x_axis_label = "Normalized capacity by test DChg divisor"
    else:
        x_axis_label = "Capacity (mAh)"

    for sample, rows in sample_to_plot_rows.items():
        if rows:
            save_sample_plot(
                sample=sample,
                sample_rows=rows,
                sample_color=color_map[sample],
                figures_dir=figures_dir,
                xlim=(
                    tuple(args.xlim)
                    if args.xlim is not None
                    else determine_sample_xlim(
                        sample_rows=rows,
                        x_min=args.x_min,
                        small_capacity_limit=args.small_capacity_limit,
                        large_capacity_limit=args.large_capacity_limit,
                    )
                ),
                ylim=tuple(args.ylim),
                dpi=args.dpi,
                show_legend=args.show_legend,
                x_axis_label=x_axis_label,
            )

    write_outputs(
        output_dir=output_dir,
        summary_rows=summary_rows,
        all_plot_rows=all_plot_rows,
        color_map=color_map,
    )

    print("\nDone.")
    print(f"Figures:      {figures_dir}")
    print(f"Plot data:    {plot_data_dir}")
    print(f"Summary file: {output_dir / 'summary' / 'stripping_summary.xlsx'}")


if __name__ == "__main__":
    main()
